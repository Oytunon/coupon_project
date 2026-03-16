from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from backend_api.app.deps import get_db, verify_api_token
from shared.models.participant import Participant
from shared.models.coupon import Coupon
from shared.models.config import SystemConfig
from shared.models.admin import AdminUser
from shared.models.event import Event
from shared.models.enrollment import EventParticipant
from backend_api.app.security import get_password_hash, get_require_full_admin
from pydantic import BaseModel, EmailStr
from datetime import datetime, timedelta

# Excel export: Türkçe sayı formatı (1.546.301,09) - binlik nokta, ondalık virgül
def _format_tr_number(n, decimals=2):
    if n is None:
        return "0,00" if decimals > 0 else "0"
    try:
        v = float(n)
        s = f"{v:.{decimals}f}"  # 1546301.09
        if "." in s:
            int_part, dec_part = s.split(".", 1)
        else:
            int_part, dec_part = s, "0" * decimals
        sign = "-" if int_part.startswith("-") else ""
        int_part = int_part.lstrip("-")
        if len(int_part) <= 3:
            return sign + int_part + ("," + dec_part if decimals > 0 else "")
        # Binlik ayırıcı: sağdan sola her 3 basamakta nokta
        result = []
        for i, c in enumerate(reversed(int_part)):
            if i > 0 and i % 3 == 0:
                result.append(".")
            result.append(c)
        int_formatted = sign + "".join(reversed(result))
        return int_formatted + ("," + dec_part if decimals > 0 else "")
    except (ValueError, TypeError):
        return str(n)

# Excel export: DB'deki tarihler UTC; Türkiye saati (UTC+3) olarak göster
def _utc_to_tr_str(dt):
    if dt is None:
        return "-"
    return (dt + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")
import openpyxl
from io import BytesIO
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/api/admin", tags=["admin"], dependencies=[Depends(verify_api_token)])


class ConfigUpdate(BaseModel):
    key: str
    value: Any




@router.get("/stats")
async def get_stats(db: Session = Depends(get_db)):
    from shared.domain.admin_stats import get_global_stats
    return get_global_stats(db)

@router.get("/participants")
async def list_participants(
    event_id: Optional[int] = None, 
    skip: int = 0,
    limit: int = 20,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    from shared.domain.participants import list_participants_paginated
    total, items = list_participants_paginated(db, event_id, skip, limit, search)
    
    return {
        "total": total,
        "items": items,
        "data_check": "paginated"
    }

@router.get("/participants/{client_id}/coupons")
async def get_user_coupons(
    client_id: int, 
    event_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 20,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    from shared.domain.participants import get_user_coupon_history
    from shared.models.coupon_event_result import CouponEventResult
    from shared.models.coupon import Coupon
    from sqlalchemy import func

    total, items = get_user_coupon_history(db, client_id, event_id, skip, limit, search)
    
    # Calculate Total Points for this context (Global or Event-Specific)
    query = db.query(func.coalesce(func.sum(CouponEventResult.points_earned), 0.0)).join(
         Coupon, Coupon.id == CouponEventResult.coupon_id
    ).filter(
        Coupon.client_id == client_id,
        CouponEventResult.is_eligible == True
    )
    
    if event_id:
        query = query.filter(CouponEventResult.event_id == event_id)
        
    total_points = query.scalar()
    
    return {
        "total": total,
        "items": items,
        "total_points": total_points
    }

@router.get("/settings")
async def get_settings(db: Session = Depends(get_db)):
    configs = db.query(SystemConfig).all()
    
    if not configs:
        # İlk çalıştırmada varsayılanları yükle
        defaults = SystemConfig.get_default_configs()
        for d in defaults:
            cfg = SystemConfig(**d)
            db.add(cfg)
        db.commit()
        configs = db.query(SystemConfig).all()
    else:
        # Mevcut ayarların açıklamalarını koddan güncelle
        defaults_map = {d["key"]: d["description"] for d in SystemConfig.get_default_configs()}
        changes_made = False
        
        for config in configs:
            if config.key in defaults_map and config.description != defaults_map[config.key]:
                config.description = defaults_map[config.key]
                changes_made = True
        
        if changes_made:
            db.commit()
    
    return {c.key: {"value": c.value, "description": c.description} for c in configs}

@router.post("/settings")
async def update_setting(update: ConfigUpdate, db: Session = Depends(get_db), _: AdminUser = Depends(get_require_full_admin)):
    config = db.query(SystemConfig).filter(SystemConfig.key == update.key).first()
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    
    config.value = update.value
    db.commit()
    return {"status": "success", "key": update.key, "new_value": update.value}

class AdminUserCreate(BaseModel):
    username: str
    password: str
    email: EmailStr
    role: str = "admin"

class AdminUserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None  # Add if we want to allow password changes
    role: Optional[str] = None
    is_active: Optional[bool] = None

class AdminUserResponse(BaseModel):
    id: int
    username: str
    email: str
    role: str
    is_active: bool

    class Config:
        from_attributes = True

@router.get("/users", response_model=List[AdminUserResponse])
async def list_admin_users(db: Session = Depends(get_db), _: AdminUser = Depends(get_require_full_admin)):
    users = db.query(AdminUser).all()
    return users

@router.post("/users", response_model=AdminUserResponse)
async def create_admin_user(user_in: AdminUserCreate, db: Session = Depends(get_db), _: AdminUser = Depends(get_require_full_admin)):
    # Username check
    existing = db.query(AdminUser).filter(AdminUser.username == user_in.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten alınmış.")
    
    # Email check
    existing_email = db.query(AdminUser).filter(AdminUser.email == user_in.email).first()
    if existing_email:
        raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kullanımda.")
    
    new_user = AdminUser(
        username=user_in.username,
        email=user_in.email,
        hashed_password=get_password_hash(user_in.password),
        role=user_in.role,
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@router.put("/users/{user_id}", response_model=AdminUserResponse)
async def update_admin_user(user_id: int, user_in: AdminUserUpdate, db: Session = Depends(get_db), _: AdminUser = Depends(get_require_full_admin)):
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")

    if user_in.username is not None and user_in.username != user.username:
        # Check uniqueness
        existing = db.query(AdminUser).filter(AdminUser.username == user_in.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten alınmış.")
        user.username = user_in.username

    if user_in.email is not None and user_in.email != user.email:
        existing_email = db.query(AdminUser).filter(AdminUser.email == user_in.email).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kullanımda.")
        user.email = user_in.email

    if user_in.password:
        user.hashed_password = get_password_hash(user_in.password)

    if user_in.role is not None:
        user.role = user_in.role

    if user_in.is_active is not None:
        user.is_active = user_in.is_active

    db.commit()
    db.refresh(user)
    return user

@router.delete("/users/{user_id}")
async def delete_admin_user(user_id: int, db: Session = Depends(get_db), current_user: AdminUser = Depends(get_require_full_admin)):
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Kendinizi silemezsiniz.")

    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")
    
    from shared.models.magic_token import MagicToken
    
    # Cascade deletions and nullifications
    db.query(Event).filter(Event.created_by == user_id).update({"created_by": None}, synchronize_session=False)
    db.query(MagicToken).filter(MagicToken.user_id == user_id).delete(synchronize_session=False)

    db.delete(user)
    db.commit()
    return {"status": "success"}

@router.get("/worker-jobs/{job_id}")
async def get_worker_job_status(job_id: int, db: Session = Depends(get_db)):
    from shared.models.worker_log import WorkerLog
    job = db.query(WorkerLog).filter(WorkerLog.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job bulunamadı")
    return {
        "id": job.id,
        "status": job.status,
        "processed": job.processed_count,
        "saved": job.saved_count,
        "total": job.total_count,
        "error": job.error_message,
        "completed_at": job.completed_at
    }

@router.post("/worker-jobs/{job_id}/cancel")
async def cancel_worker_job(job_id: int, db: Session = Depends(get_db), _: AdminUser = Depends(get_require_full_admin)):
    from shared.models.worker_log import WorkerLog
    job = db.query(WorkerLog).filter(WorkerLog.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job bulunamadı")
    
    if job.status not in ["pending", "running"]:
        return {"status": "error", "message": "Zaten tamamlanmış veya durdurulmuş bir işlem iptal edilemez."}
    
    job.status = "cancelled"
    job.completed_at = datetime.utcnow()
    db.commit()
    return {"status": "success", "message": "İşlem iptal ediliyor..."}

@router.get("/events/{event_id}/stats")
async def get_event_stats_api(event_id: int, db: Session = Depends(get_db)):
    from shared.models.enrollment import EventParticipant
    from shared.models.coupon_event_result import CouponEventResult
    from sqlalchemy import func
    from shared.models.event import Event
    
    total_participants = db.query(EventParticipant).filter(EventParticipant.event_id == event_id).count()
    
    # Calculate total points (Live from results for accuracy)
    points = db.query(func.coalesce(func.sum(CouponEventResult.points_earned), 0.0)).filter(
        CouponEventResult.event_id == event_id,
        CouponEventResult.is_eligible == True
    ).scalar()
    
    # Calculate coupon stats (Live from results)
    total_coupons = db.query(func.count(CouponEventResult.id)).filter(
        CouponEventResult.event_id == event_id,
        CouponEventResult.is_eligible == True
    ).scalar()
    
    # Calculate total stake (Live from results joined with Coupon)
    total_stake = db.query(func.coalesce(func.sum(Coupon.stake), 0.0)).join(
        CouponEventResult, Coupon.id == CouponEventResult.coupon_id
    ).filter(
        CouponEventResult.event_id == event_id,
        CouponEventResult.is_eligible == True
    ).scalar()

    return {
        "event_id": event_id,
        "total_participants": total_participants,
        "total_points_distributed": float(points),
        "total_coupons": total_coupons,
        "total_stake": float(total_stake),
        "event_name": db.query(Event.name).filter(Event.id == event_id).scalar() or "Unknown",
        "status": db.query(Event.status).filter(Event.id == event_id).scalar() or "draft"
    }


@router.get("/events/{event_id}/statistics")
async def get_event_statistics_api(event_id: int, db: Session = Depends(get_db)):
    """
    Turnuva istatistikleri: KATILIM, TOPLAM, KAZANAN, KAYBEDEN, YATIRIM metrikleri.
    """
    from shared.models.coupon_event_result import CouponEventResult
    from sqlalchemy import func
    from sqlalchemy import or_

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event bulunamadı")

    # KATILIM
    total_participants = db.query(EventParticipant).filter(EventParticipant.event_id == event_id).count()
    participants_with_points = db.query(EventParticipant).filter(
        EventParticipant.event_id == event_id, EventParticipant.total_points > 0
    ).count()
    participants_without_points = total_participants - participants_with_points

    # CouponEventResult üzerinden event-kupon eşleşmesi (multi-event desteği)
    cer_subq = db.query(CouponEventResult.coupon_id).filter(CouponEventResult.event_id == event_id)

    # TOPLAM (Kazanan + Kaybeden kuponlar)
    won_lost_subq = db.query(Coupon.id).filter(
        Coupon.id.in_(cer_subq),
        or_(
            or_(Coupon.state == "won", Coupon.state == "Won"),
            or_(Coupon.state == "lost", Coupon.state == "Lost")
        )
    )
    total_coupons = db.query(func.count(Coupon.id)).filter(Coupon.id.in_(won_lost_subq)).scalar() or 0
    total_stake = db.query(func.coalesce(func.sum(Coupon.stake), 0.0)).filter(Coupon.id.in_(won_lost_subq)).scalar() or 0.0

    # KAZANAN (Won)
    won_coupons = db.query(func.count(Coupon.id)).filter(
        Coupon.id.in_(cer_subq),
        or_(Coupon.state == "won", Coupon.state == "Won")
    ).scalar() or 0
    won_stake = db.query(func.coalesce(func.sum(Coupon.stake), 0.0)).filter(
        Coupon.id.in_(cer_subq),
        or_(Coupon.state == "won", Coupon.state == "Won")
    ).scalar() or 0.0
    won_payout = db.query(func.coalesce(func.sum(Coupon.winning), 0.0)).filter(
        Coupon.id.in_(cer_subq),
        or_(Coupon.state == "won", Coupon.state == "Won")
    ).scalar() or 0.0

    # KAYBEDEN / GEÇERLİ OLMAYAN (Lost veya is_eligible=False)
    lost_or_ineligible_subq = (
        db.query(CouponEventResult.coupon_id)
        .join(Coupon, Coupon.id == CouponEventResult.coupon_id)
        .filter(
            CouponEventResult.event_id == event_id,
            or_(
                or_(Coupon.state == "lost", Coupon.state == "Lost"),
                CouponEventResult.is_eligible == False
            )
        )
    )
    lost_coupons = db.query(func.count(Coupon.id)).filter(
        Coupon.id.in_(lost_or_ineligible_subq)
    ).scalar() or 0
    lost_stake = db.query(func.coalesce(func.sum(Coupon.stake), 0.0)).join(
        CouponEventResult, Coupon.id == CouponEventResult.coupon_id
    ).filter(
        CouponEventResult.event_id == event_id,
        or_(
            or_(Coupon.state == "lost", Coupon.state == "Lost"),
            CouponEventResult.is_eligible == False
        )
    ).scalar() or 0.0

    # YATIRIM (lig uyumlu kazanan + kaybeden kuponların stake toplamı - kupona basılan mebla)
    eligible_cer_subq = db.query(CouponEventResult.coupon_id).filter(
        CouponEventResult.event_id == event_id,
        CouponEventResult.is_eligible == True
    )
    total_investment = db.query(func.coalesce(func.sum(Coupon.stake), 0.0)).filter(
        Coupon.id.in_(eligible_cer_subq),
        or_(Coupon.state == "won", Coupon.state == "Won", Coupon.state == "lost", Coupon.state == "Lost")
    ).scalar() or 0.0

    return {
        "event_id": event_id,
        "event_name": event.name,
        "status": event.status or "draft",
        "katilim": {
            "total_participants": total_participants,
            "participants_with_points": participants_with_points,
            "participants_without_points": participants_without_points,
        },
        "toplam": {
            "total_coupons": total_coupons,
            "total_stake": float(total_stake),
        },
        "kazanan": {
            "won_coupons": won_coupons,
            "won_stake": float(won_stake),
            "won_payout": float(won_payout),
        },
        "kaybeden": {
            "lost_coupons": lost_coupons,
            "lost_stake": float(lost_stake),
        },
        "yatirim": {
            "total_investment": float(total_investment),
        },
    }


@router.get("/events/{event_id}/statistics/export")
async def export_event_statistics(event_id: int, db: Session = Depends(get_db)):
    """Turnuva istatistiklerini Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event bulunamadı")

    stats = await get_event_statistics_api(event_id, db)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Istatistikler"
    ws1.append(["Metrik", "Deger"])
    ws1.append(["Toplam Katilimci Sayisi", _format_tr_number(stats["katilim"]["total_participants"], 0)])
    ws1.append(["Toplam Kupon Sayisi", _format_tr_number(stats["toplam"]["total_coupons"], 0)])
    ws1.append(["Toplam Bahis Tutari (TL)", _format_tr_number(stats["toplam"]["total_stake"])])
    ws1.append(["Gecerli Kupon Sayisi", _format_tr_number(stats["kazanan"]["won_coupons"], 0)])
    ws1.append(["Gecerli Kuponlarin Bahis Tutari (TL)", _format_tr_number(stats["kazanan"]["won_stake"])])
    ws1.append(["Gecerli Kuponlarin Kazanc Tutari (TL)", _format_tr_number(stats["kazanan"]["won_payout"])])
    ws1.append(["Gecerli Olmayan Kupon Sayisi", _format_tr_number(stats["kaybeden"]["lost_coupons"], 0)])
    ws1.append(["Gecerli Olmayan Kuponlarin Bahis Tutari (TL)", _format_tr_number(stats["kaybeden"]["lost_stake"])])
    ws1.append(["Katilimcilarin Toplam Yatirim Tutari (TL)", _format_tr_number(stats["yatirim"]["total_investment"])])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.slug}_istatistikler_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/events/{event_id}/lost-coupons")
async def get_event_lost_coupons_api(
    event_id: int,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """
    Lig uyumlu kaybeden kuponlar listesi (event_lost_coupons tablosundan).
    """
    from shared.models.event_lost_coupon import EventLostCoupon

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event bulunamadı")

    q = (
        db.query(EventLostCoupon, Participant.username)
        .outerjoin(Participant, Participant.client_id == EventLostCoupon.client_id)
        .filter(EventLostCoupon.event_id == event_id)
        .order_by(EventLostCoupon.created_at.desc())
    )
    total = db.query(EventLostCoupon).filter(EventLostCoupon.event_id == event_id).count()
    rows = q.offset(skip).limit(limit).all()

    items = []
    for elc, username in rows:
        items.append({
            "id": elc.id,
            "coupon_id": elc.coupon_id,
            "username": username or "-",
            "stake": float(elc.stake),
            "created_at": elc.created_at.isoformat() if elc.created_at else None,
        })

    return {
        "event_id": event_id,
        "total": total,
        "items": items,
    }


@router.get("/events/{event_id}/participants")
async def get_event_participants_api(event_id: int, db: Session = Depends(get_db)):
    """
    Detailed participant list for a specific event.
    """
    from shared.models.enrollment import EventParticipant
    from shared.models.participant import Participant
    from shared.models.coupon_event_result import CouponEventResult
    from shared.models.coupon import Coupon
    from sqlalchemy import func
    
    # Fetch basic participant info
    results = (
        db.query(EventParticipant, Participant.username, Participant.client_id)
        .join(Participant, EventParticipant.participant_id == Participant.id)
        .filter(EventParticipant.event_id == event_id)
        .all()
    )
    
    items = []
    for ep, uname, cid in results:
        # Get count of coupons via CouponEventResult
        c_count = db.query(func.count(CouponEventResult.id)).join(
             Coupon, Coupon.id == CouponEventResult.coupon_id
        ).filter(
            CouponEventResult.event_id == event_id,
            Coupon.client_id == cid,
            CouponEventResult.is_eligible == True
        ).scalar()
        
        # Get live points via CouponEventResult (Fix for discrepancy)
        live_points = db.query(func.coalesce(func.sum(CouponEventResult.points_earned), 0.0)).join(
             Coupon, Coupon.id == CouponEventResult.coupon_id
        ).filter(
            CouponEventResult.event_id == event_id,
            Coupon.client_id == cid,
            CouponEventResult.is_eligible == True
        ).scalar()
        
        items.append({
            "id": ep.participant_id,
            "username": uname,
            "client_id": cid,
            "points": live_points, # Use live points
            "coupon_count": c_count,
            "joined_at": ep.joined_at
        })
    
    # Sort by points in memory (since we calculate live)
    items.sort(key=lambda x: x["points"], reverse=True)
        
    return {
         "total": len(items),
         "items": items
    }

@router.get("/events/{event_id}/participants/export")
async def export_event_participants(event_id: int, db: Session = Depends(get_db)):
    """Etkinlik katılımcılarını Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
    
    from shared.domain.leaderboard import get_event_leaderboard
    results = get_event_leaderboard(db, event_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Katilimcilar"
    ws.append(["ID", "Kullanici Adi", "Client ID", "Katilim Tarihi", "Kupon Sayisi", "Toplam Puan"])
    
    for p in results:
        ws.append([
            p["id"], p["username"], p["client_id"],
            _utc_to_tr_str(p["joined_at"]),
            p["coupon_count"], p["points"]
        ])
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{event.slug}_participants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/events/{event_id}/coupons/export")
async def export_event_coupons(event_id: int, db: Session = Depends(get_db)):
    """Etkinlikte puan kazanan kuponları Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")

    from shared.models.coupon_event_result import CouponEventResult
    
    # Sadece puan kazandıran kuponları çek (is_processed=True)
    # FIX: Use CouponEventResult to get accurate event-specific data
    results = db.query(Coupon, CouponEventResult).join(
        CouponEventResult, CouponEventResult.coupon_id == Coupon.id
    ).filter(
        CouponEventResult.event_id == event_id,
        Coupon.is_processed == True,
        CouponEventResult.is_eligible == True
    ).order_by(Coupon.processed_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kuponlar"
    ws.append(["ID", "Client ID", "Bet ID", "Tarih", "Stake", "Odds", "State", "Puan", "Live?"])

    for c, cer in results:
        ws.append([
            c.id, c.client_id, c.bet_id,
            _utc_to_tr_str(c.created_at),
            c.stake, c.odds, c.state, 
            cer.points_earned, # Use event-specific points
            "YES" if c.is_live else "NO"
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.slug}_coupons_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@router.get("/events/{event_id}/reward-history")
async def get_event_reward_history(
    event_id: int,
    db: Session = Depends(get_db)
):
    from shared.models.reward_job import RewardJob
    from shared.models.participant import Participant
    
    # Fetch all reward jobs for this event
    jobs = db.query(RewardJob).filter(RewardJob.event_id == event_id).order_by(RewardJob.created_at.desc()).all()
    
    all_rewards = []
    
    for job in jobs:
        results = job.results or {}
        for client_id_str, rewards_list in results.items():
            try:
                client_id = int(client_id_str)
            except ValueError:
                continue
                
            username = db.query(Participant.username).filter(Participant.client_id == client_id).scalar() or "Unknown"
            
            for r in rewards_list:
                if r.get("status") == "success":
                    all_rewards.append({
                        "job_id": job.id,
                        "client_id": client_id,
                        "username": username,
                        "reward_type": r.get("rule", {}).get("reward_type"),
                        "amount": r.get("rule", {}).get("amount"),
                        "timestamp": r.get("timestamp"),
                        "status": "success"
                    })
    
    return all_rewards
