from typing import List, Optional, Dict, Any
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel, Field
from io import BytesIO
import openpyxl
import uuid
import shutil
import os

from shared.database import get_db_session
from shared.models.event import Event
from shared.models.coupon_event_result import CouponEventResult
from shared.models.admin import AdminUser
from shared.models.coupon import Coupon
from shared.models.participant import Participant
from shared.models.enrollment import EventParticipant
from backend_api.app.security import get_current_admin, get_require_full_admin

router = APIRouter(prefix="/api/admin/events", tags=["admin-events"])


# === Pydantic Schemas ===


class ContentRule(BaseModel):
    title: str
    content: str
    icon: str = "Info"


class RewardRule(BaseModel):
    reward_type: str  # 'cash', 'bonus', 'freebet', 'spin'
    amount: float
    currency: str = "TRY"
    criteria_type: str # 'rank', 'min_points'
    criteria_value: int # e.g. 10 (rank <= 10) or 5000 (points >= 5000)
    partner_bonus_id: Optional[int] = None


class EventRules(BaseModel):
    model_config = {"extra": "allow"}
    min_stake: float = 100.0
    min_odd: float = 1.5
    min_combination: int = 2
    max_combination: Optional[int] = None
    allowed_league_ids: List[int] = []
    max_coupons_per_user: Optional[int] = None
    scoring_formula: str = "stake_times_odds"  # 'simple', 'stake_times_odds', 'combo_bonus'
    combo_bonus_enabled: bool = False
    combo_bonus_multiplier: float = 0.1
    min_deposit: int = 1000
    prize_pool: float = 0.0
    rewards: List[RewardRule] = []


class EventCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    slug: Optional[str] = Field(None, min_length=1, max_length=50)
    description: Optional[str] = None
    start_date: datetime
    end_date: datetime
    display_until: Optional[datetime] = None
    won_point_multiplier: float = 1.0
    loss_point_multiplier: float = 0.0
    image_url: Optional[str] = None
    rules: EventRules
    content_rules: List[ContentRule] = []


class EventUpdate(BaseModel):
    model_config = {"extra": "allow"}
    name: Optional[str] = Field(None, min_length=1, max_length=255)
    slug: Optional[str] = Field(None, min_length=1, max_length=50)
    description: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    display_until: Optional[datetime] = None
    won_point_multiplier: Optional[float] = None
    loss_point_multiplier: Optional[float] = None
    image_url: Optional[str] = None
    rules: Optional[EventRules] = None
    content_rules: Optional[List[ContentRule]] = None


class EventStatusUpdate(BaseModel):
    status: str = Field(..., pattern="^(draft|active|paused|ended|archived)$")


class EventResponse(BaseModel):
    id: int
    name: str
    slug: str
    description: Optional[str]
    status: str
    start_date: datetime
    end_date: datetime
    display_until: Optional[datetime] = None
    won_point_multiplier: float
    loss_point_multiplier: float
    image_url: Optional[str]
    rules: dict
    content_rules: List[ContentRule] = []
    reward_status: Optional[str] = None
    last_worker_run: Optional[datetime] = None
    last_cron_run: Optional[datetime] = None
    avg_worker_duration_seconds: Optional[float] = None
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True


class EventStatsResponse(BaseModel):
    event_id: int
    event_name: str
    status: str
    total_participants: int
    total_coupons: int
    eligible_coupons: int
    won_coupons: int
    lost_coupons: int
    pending_coupons: int
    total_stake: float
    total_points_distributed: float
    avg_points_per_user: float
    avg_stake: float



# === Endpoints ===

@router.post("", response_model=EventResponse, status_code=201)
async def create_event(
    event_data: EventCreate,
    db: Session = Depends(get_db_session),
    current_admin: AdminUser = Depends(get_require_full_admin)
):
    """Yeni event oluştur."""
    try:
        # Date validation
        if event_data.end_date <= event_data.start_date:
            raise HTTPException(400, "End date must be after start date")
        
        # Auto-generate slug if not provided
        slug_to_use = event_data.slug
        if not slug_to_use:
            from slugify import slugify
            import uuid
            base_slug = slugify(event_data.name)
            slug_to_use = f"{base_slug}-{uuid.uuid4().hex[:6]}"

        # Slug uniqueness ensure
        existing = db.query(Event).filter(Event.slug == slug_to_use).first()
        if existing:
             # Retry once with different random suffix if collision (rare)
             import uuid
             from slugify import slugify
             base_slug = slugify(event_data.name)
             slug_to_use = f"{base_slug}-{uuid.uuid4().hex[:6]}"

        processed_content_rules = []
        if event_data.content_rules:
            for rule in event_data.content_rules:
                if hasattr(rule, "model_dump"):
                    processed_content_rules.append(rule.model_dump())
                elif hasattr(rule, "dict"):
                    processed_content_rules.append(rule.dict())
                else:
                    processed_content_rules.append(rule)

        # stake_times_odds_raw: çarpanlar sabit (kazanan=1, kaybeden=0)
        scoring_formula = getattr(event_data.rules, 'scoring_formula', None)
        if scoring_formula == "stake_times_odds_raw":
            won_mult, loss_mult = 1.0, 0.0
        else:
            won_mult, loss_mult = event_data.won_point_multiplier, event_data.loss_point_multiplier

        event = Event(
            name=event_data.name,
            slug=slug_to_use,
            description=event_data.description,
            start_date=event_data.start_date,
            end_date=event_data.end_date,
            display_until=event_data.display_until,
            won_point_multiplier=won_mult,
            loss_point_multiplier=loss_mult,
            image_url=event_data.image_url,
            rules=event_data.rules.dict(),
            content_rules=processed_content_rules,
            created_by=current_admin.id,
            status="draft" 
        )
        
        db.add(event)
        db.commit()
        db.refresh(event)
        
        return event
    except Exception as e:
        raise e


@router.get("", response_model=List[EventResponse])
async def list_events(
    status: Optional[str] = Query(None, pattern="^(draft|active|paused|ended|archived)$"),
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Tüm eventleri listele."""
    query = db.query(Event)
    if status:
        query = query.filter(Event.status == status)
        
    events = query.order_by(Event.created_at.desc()).all()
    
    # Her event için son ödül dağıtım durumunu toplu çek (Subquery optimization)
    from shared.models.reward_job import RewardJob
    from sqlalchemy import select, func
    
    # Her event için en son job id'sini bul
    subquery = db.query(
        RewardJob.event_id,
        func.max(RewardJob.id).label('max_job_id')
    ).group_by(RewardJob.event_id).subquery()
    
    # Bu id'lere karşılık gelen status'ları çek
    job_statuses = {
        row.event_id: row.status 
        for row in db.query(RewardJob.event_id, RewardJob.status)
        .join(subquery, RewardJob.id == subquery.c.max_job_id)
        .all()
    }
    
    for event in events:
        event.reward_status = job_statuses.get(event.id, "none")
    
    # Her event için son worker çalışma tarihini bul (manuel vs otomatik ayrımı)
    from shared.models.worker_log import WorkerLog
    
    # Manuel çalıştırmalar (event_id belirtilmiş olanlar)
    manual_subquery = db.query(
        WorkerLog.event_id,
        func.max(WorkerLog.completed_at).label('last_run')
    ).filter(
        WorkerLog.status == "completed",
        WorkerLog.event_id.isnot(None)
    ).group_by(WorkerLog.event_id).subquery()
    
    last_manual_runs = {
        row.event_id: row.last_run
        for row in db.query(manual_subquery.c.event_id, manual_subquery.c.last_run).all()
    }
    
    # Otomatik (cron) çalıştırmalar (event_id=None)
    cron_last_run = db.query(func.max(WorkerLog.completed_at)).filter(
        WorkerLog.status == "completed",
        WorkerLog.event_id.is_(None)
    ).scalar()
    
    # Her event için son 5 tamamlanmış worker'ın ortalama süresi (geri sayım için)
    event_ids = [e.id for e in events]
    completed_logs = db.query(WorkerLog).filter(
        WorkerLog.event_id.in_(event_ids),
        WorkerLog.status == "completed",
        WorkerLog.completed_at.isnot(None),
        WorkerLog.created_at.isnot(None)
    ).order_by(WorkerLog.event_id, WorkerLog.completed_at.desc()).all()
    from collections import defaultdict
    by_event = defaultdict(list)
    for log in completed_logs:
        if len(by_event[log.event_id]) < 5:
            by_event[log.event_id].append(log)
    avg_durations = {}
    for eid, logs in by_event.items():
        durations = [(l.completed_at - l.created_at).total_seconds() for l in logs]
        if durations:
            avg_durations[eid] = round(sum(durations) / len(durations), 1)

    for event in events:
        event.last_worker_run = last_manual_runs.get(event.id)
        event.last_cron_run = cron_last_run
        event.avg_worker_duration_seconds = avg_durations.get(event.id)
            
    return events


@router.get("/{event_id}", response_model=EventResponse)
async def get_event(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Event detayını getir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
        
    from shared.models.reward_job import RewardJob
    last_job = db.query(RewardJob).filter(RewardJob.event_id == event.id).order_by(RewardJob.created_at.desc()).first()
    event.reward_status = last_job.status if last_job else "none"
    
    return event


@router.put("/{event_id}", response_model=EventResponse)
async def update_event(
    event_id: int,
    event_data: EventUpdate,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_require_full_admin)
):
    """Event'i güncelle."""
    import logging
    from sqlalchemy.orm.attributes import flag_modified
    logger = logging.getLogger("backend_api")
    
    # Received data log
    data_dict = event_data.dict(exclude_unset=True)
    logger.info(f"PUT /admin/events/{event_id} - Body: {data_dict}")

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        logger.error(f"Event not found: {event_id}")
        raise HTTPException(404, "Event not found")
    
    # 1. Slug check
    if "slug" in data_dict and data_dict["slug"] != event.slug:
        existing = db.query(Event).filter(Event.slug == data_dict["slug"]).first()
        if existing:
            logger.error(f"Slug already in use: {data_dict['slug']}")
            raise HTTPException(400, "Slug already in use")

    # 2. Update simple fields
    for field, value in data_dict.items():
        if field not in ["rules", "content_rules"]:
            logger.info(f"Updating {field}: {getattr(event, field)} -> {value}")
            setattr(event, field, value)

    # 3. Update rules (Nested merge)
    if "rules" in data_dict and data_dict["rules"]:
        current_rules = dict(event.rules or {})
        new_rules = data_dict["rules"]
        if hasattr(new_rules, "dict"):
            new_rules = new_rules.dict(exclude_unset=True)
        elif hasattr(new_rules, "model_dump"):
            new_rules = new_rules.model_dump(exclude_unset=True)
            
        merged_rules = {**current_rules, **new_rules}
        logger.info(f"Rules update: {current_rules} -> {merged_rules}")
        event.rules = merged_rules
        # stake_times_odds_raw: çarpanlar sabit (kazanan=1, kaybeden=0)
        if merged_rules.get("scoring_formula") == "stake_times_odds_raw":
            event.won_point_multiplier = 1.0
            event.loss_point_multiplier = 0.0
        # Explicitly mark rules as modified for SQLAlchemy
        flag_modified(event, "rules")

    # 4. Update content_rules (Full replace)
    if "content_rules" in data_dict:
        new_content_rules = data_dict["content_rules"]
        # Convert Pydantic models to dicts if necessary
        processed_rules = []
        for rule in new_content_rules:
            if hasattr(rule, "model_dump"):
                processed_rules.append(rule.model_dump())
            elif hasattr(rule, "dict"):
                processed_rules.append(rule.dict())
            else:
                processed_rules.append(rule)
        
        logger.info(f"Content Rules update: {event.content_rules} -> {processed_rules}")
        event.content_rules = processed_rules
        flag_modified(event, "content_rules")
    
    # 5. Final validation
    if event.end_date <= event.start_date:
        logger.error(f"Invalid dates: {event.start_date} -> {event.end_date}")
        raise HTTPException(400, "End date must be after start date")
    
    try:
        event.updated_at = datetime.now()
        db.commit()
        db.refresh(event)
        logger.info(f"Event {event_id} successfully saved and committed")
        return event
    except Exception as e:
        db.rollback()
        logger.error(f"Save failed for event {event_id}: {str(e)}")
        raise HTTPException(500, f"Database save error: {str(e)}")


@router.delete("/{event_id}", status_code=204)
async def delete_event(
    event_id: int,
    db: Session = Depends(get_db_session),
    current_admin: AdminUser = Depends(get_require_full_admin)
):
    """Event'i sil."""
    # Updated: Allow regular admins to delete events as well, since there is no UI to create superadmins easily yet.
    if current_admin.role not in ["admin", "superadmin"]:
        raise HTTPException(403, "Not authorized to delete events")
    
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
    
    try:
        # Manually delete related records to bypass missing DB-level CASCADE
        from shared.models.worker_log import WorkerLog
        from sqlalchemy import exists, and_, tuple_

        # 0. Identify coupons involved in this event (for potential cleanup)
        # We want to delete coupons that become 'orphans' (no longer attached to any event)
        # But we must ensure we don't delete coupons attached to OTHER events.
        
        # Get IDs of coupons that have a result for this event
        involved_coupon_ids = [
            cid for (cid,) in db.query(CouponEventResult.coupon_id)
            .filter(CouponEventResult.event_id == event_id)
            .all()
        ]
        
        # Also consider coupons primarily linked via event_id (if any, though usually redundant)
        primary_coupon_ids = [
            cid for (cid,) in db.query(Coupon.id)
            .filter(Coupon.event_id == event_id)
            .all()
        ]
        
        all_candidate_ids = set(involved_coupon_ids + primary_coupon_ids)
        
        # 1. Delete Worker Logs
        db.query(WorkerLog).filter(WorkerLog.event_id == event_id).delete(synchronize_session=False)
        
        # 2. Delete Results (This removes the 'link' to this event)
        db.query(CouponEventResult).filter(CouponEventResult.event_id == event_id).delete(synchronize_session=False)

        # 3. Delete Participants (Enrollment)
        db.query(EventParticipant).filter(EventParticipant.event_id == event_id).delete(synchronize_session=False)

        # 4. Smart Coupon Cleanup
        # For the candidate coupons, check if they are still used by ANY other event
        # definition of 'used': Has a CouponEventResult for another event
        
        # We can implement this as a bulk delete with subquery logic, 
        # or iterate if the list is small. A query is safer/faster.
        
        if all_candidate_ids:
            # Delete coupons that:
            # 1. Are in our candidate list (involved in the deleted event)
            # 2. Do NOT have any remaining CouponEventResult rows
            
            # Note: We already deleted the results for THIS event in step 2.
            # So looking for CouponEventResult now effectively checks for "Other Events".
            
            orphaned_coupons_query = db.query(Coupon.id).filter(
                Coupon.id.in_(all_candidate_ids),
                ~exists().where(CouponEventResult.coupon_id == Coupon.id)
            )
            
            # Execute delete
            deleted_count = db.query(Coupon).filter(
                Coupon.id.in_(orphaned_coupons_query)
            ).delete(synchronize_session=False)
            
            print(f"DEBUG: Deleted {deleted_count} orphan coupons for event {event_id}")

        # 5. For any remaining coupons that were primarily linked to this event but NOT deleted 
        # (because they are shared), we must set event_id to Null to avoid FK issues if we didn't delete them.
        db.query(Coupon).filter(Coupon.event_id == event_id).update({"event_id": None}, synchronize_session=False)
        
        # Now delete event
        db.delete(event)
        db.commit()
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        # Return the specific error to the client
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")
        
    return None


@router.patch("/{event_id}/status", response_model=EventResponse)
async def update_event_status(
    event_id: int,
    status_data: EventStatusUpdate,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_require_full_admin)
):
    """Event status'ünü güncelle."""
    import logging
    logger = logging.getLogger("backend_api")
    logger.info(f"PATCH /admin/events/{event_id}/status: {status_data.status}")

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        logger.error(f"Event not found: {event_id}")
        raise HTTPException(404, "Event not found")
    
    event.status = status_data.status
    event.updated_at = datetime.now()
    db.commit()
    db.refresh(event)
    logger.info(f"Status updated successfully for event {event_id}")
    return event


@router.get("/{event_id}/stats", response_model=EventStatsResponse)
async def get_event_stats(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Event istatistiklerini getir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
    
    from shared.domain.event_stats import get_event_metrics
    metrics = get_event_metrics(db, event_id)
    
    return EventStatsResponse(
        event_id=event_id,
        event_name=event.name,
        status=event.status,
        **metrics
    )


@router.post("/{event_id}/recalculate", status_code=202)
async def recalculate_event_points(
    event_id: int,
    db: Session = Depends(get_db_session),
    current_admin: AdminUser = Depends(get_require_full_admin)
):
    """Event'in tüm puanlarını yeniden hesapla."""
    
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
    
    from shared.domain.scoring_engine import calculate_points_for_event
    
    results = db.query(CouponEventResult).filter(
        CouponEventResult.event_id == event_id,
        CouponEventResult.is_eligible == True
    ).all()
    
    recalculated_count = 0
    for cer in results:
        coupon = db.query(Coupon).filter(Coupon.id == cer.coupon_id).first()
        if coupon:
            new_points, calculation = calculate_points_for_event(coupon, event)
            cer.points_earned = new_points
            cer.points_calculation = calculation
            cer.evaluated_at = datetime.now()
            recalculated_count += 1

    # Katılımcı toplam puanlarını güncelle (sıralama için)
    enrollments = db.query(EventParticipant).filter(EventParticipant.event_id == event_id).all()
    for ep in enrollments:
        participant = db.query(Participant).filter(Participant.id == ep.participant_id).first()
        if participant:
            total = db.query(func.sum(CouponEventResult.points_earned)).filter(
                CouponEventResult.event_id == event_id,
                CouponEventResult.coupon_id.in_(db.query(Coupon.id).filter(Coupon.client_id == participant.client_id)),
                CouponEventResult.is_eligible == True
            ).scalar() or 0.0
            ep.total_points = total

    db.commit()
    return {
        "message": f"Recalculated {recalculated_count} coupon points for event {event.name}",
        "recalculated_count": recalculated_count
    }


@router.post("/{event_id}/worker", status_code=202)
async def run_event_worker(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_require_full_admin)
):
    """Worker'ı manuel tetikle."""
    from shared.models.worker_log import WorkerLog
    from shared.domain.scoring_engine import process_coupons
    
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
    
    # Eşzamanlılık kilidi: Başka bir worker çalışıyor mu?
    running_job = db.query(WorkerLog).filter(WorkerLog.status.in_(["running", "pending"])).first()
    if running_job:
        raise HTTPException(
            status_code=409,
            detail={
                "message": f"Başka bir worker zaten çalışıyor veya bekliyor (Job ID: {running_job.id}). Lütfen tamamlanmasını bekleyin.",
                "running_job_id": running_job.id,
            }
        )
    
    job = WorkerLog(event_id=event_id, status="pending")
    db.add(job)
    db.commit()
    db.refresh(job)

    import asyncio
    asyncio.create_task(process_coupons(target_event_id=event_id, job_id=job.id, scan_hours=42))
    return {"status": "initiated", "message": "Worker started", "job_id": job.id}


@router.post("/{event_id}/distribute-rewards", status_code=202)
async def distribute_event_rewards(
    event_id: int,
    db: Session = Depends(get_db_session),
    current_admin: AdminUser = Depends(get_require_full_admin)
):
    """Event ödüllerini dağıtmak için worker'ı tetikle."""
    from shared.models.reward_job import RewardJob
    
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
        
    # Strict Check: Event MUST be ended
    if event.status != 'ended':
        raise HTTPException(400, "Ödül dağıtımı için etkinlik bitmiş olmalıdır (Status: Ended).")
    
    # Create Job
    job = RewardJob(event_id=event_id, event_name_snapshot=event.name, status="pending")
    db.add(job)
    db.commit()
    db.refresh(job)
    
    # We rely on the standalone worker to pick this up using the scheduler or polling.
    # If we want to trigger it immediately, we might need a way to signal it.
    # But user asked for "separate worker", implying async processing.
    
    return {"status": "queued", "message": "Reward distribution queued", "job_id": job.id}


@router.get("/{event_id}/participants")
async def get_event_participants_list(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Etkinliğe katılan kullanıcıları getir (Puan sıralı)."""
    from shared.domain.leaderboard import get_event_leaderboard
    return get_event_leaderboard(db, event_id)


@router.get("/{event_id}/participants/export")
async def export_event_participants(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Etkinlik katılımcılarını Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")

    from shared.domain.leaderboard import get_event_leaderboard
    results = get_event_leaderboard(db, event_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Katilimcilar"
    ws.append(["ID", "Kullanici Adi", "Client ID", "Katilim Tarihi", "Kupon Sayisi", "Toplam Puan"])

    for p in results:
        ws.append([
            p["id"], p["username"], p["client_id"],
            p["joined_at"].strftime("%Y-%m-%d %H:%M:%S") if p["joined_at"] else "-",
            p["coupon_count"], p["points"]
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.slug}_participants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get("/{event_id}/coupons/export")
async def export_event_coupons(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Etkinlikte puan kazanan kuponları Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")

    # Sadece puan kazandıran kuponları çek (is_processed=True ve calculation > 0)
    coupons = db.query(Coupon).filter(
        Coupon.event_id == event_id,
        Coupon.is_processed == True
    ).order_by(Coupon.processed_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kuponlar"
    ws.append(["ID", "Client ID", "Bet ID", "Tarih", "Stake", "Odds", "State", "Puan", "Live?"])

    for c in coupons:
        ws.append([
            c.id, c.client_id, c.bet_id,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "-",
            c.stake, c.odds, c.state, c.calculation,
            "YES" if c.is_live else "NO"
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.slug}_coupons_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get("/{event_id}/reward-history/export")
async def export_reward_history(
    event_id: int,
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_current_admin)
):
    """Dağıtılan ödüllerin listesini Excel olarak indir."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")

    from shared.models.reward_job import RewardJob
    from shared.models.participant import Participant

    # Fetch all successful rewards for this event
    jobs = db.query(RewardJob).filter(RewardJob.event_id == event_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dagitilan Oduller"
    ws.append(["Tarih", "Kullanici Adi", "Client ID", "Odul Turu", "Miktar", "Durum"])

    for job in jobs:
        results = job.results or {}
        if not isinstance(results, dict):
            continue
            
        for client_id_str, rewards_list in results.items():
            if not isinstance(rewards_list, list):
                continue
                
            # Try to get username, handle non-integer client_id_str safely
            try:
                c_id = int(client_id_str)
                username = db.query(Participant.username).filter(Participant.client_id == c_id).scalar() or f"User-{client_id_str}"
            except (ValueError, TypeError):
                username = f"User-{client_id_str}"
                
            for r in rewards_list:
                if not isinstance(r, dict):
                    continue
                    
                if r.get("status") == "success":
                    # Robust timestamp parsing
                    ts_raw = r.get("timestamp", "")
                    ts_display = ts_raw.split(".")[0].replace("T", " ") if ts_raw else "-"
                    
                    rule = r.get("rule", {})
                    ws.append([
                        ts_display,
                        username,
                        client_id_str,
                        rule.get("reward_type", "") if isinstance(rule, dict) else "",
                        rule.get("amount", 0) if isinstance(rule, dict) else 0,
                        "Basarili"
                    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{event.slug}_rewards_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition"
        },
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.post("/{event_id}/upload-image", response_model=Dict[str, str])
async def upload_event_image(
    event_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db_session),
    _: AdminUser = Depends(get_require_full_admin)
):
    """Event için görsel yükle."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Event not found")
        
    # Validation: Only images
    if not file.content_type.startswith("image/"):
        raise HTTPException(400, "Sadece resim dosyaları yüklenebilir.")
        
    # Create filename
    ext = os.path.splitext(file.filename)[1]
    filename = f"event_{event_id}_{uuid.uuid4().hex}{ext}"
    
    # Ensure directory exists
    upload_dir = "media"
    os.makedirs(upload_dir, exist_ok=True)
    
    file_path = os.path.join(upload_dir, filename)
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Update event model
        image_url = f"/media/{filename}"
        event.image_url = image_url
        db.commit()
        
        return {"image_url": image_url}
    except Exception as e:
        import logging
        logger = logging.getLogger("backend_api")
        logger.error(f"Image upload failed: {str(e)}")
        raise HTTPException(500, f"Görsel yükleme başarısız oldu: {str(e)}")

